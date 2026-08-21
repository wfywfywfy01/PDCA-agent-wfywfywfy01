import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

monthly = [
    {"month": "2026-01", "performance_cny": 618096.15432,       "line_count": 22},
    {"month": "2026-02", "performance_cny": 285708.36798,       "line_count": 23},
    {"month": "2026-03", "performance_cny": 17236.38744,        "line_count": 9},
    {"month": "2026-04", "performance_cny": 580265.910125,      "line_count": 40},
    {"month": "2026-05", "performance_cny": 515171.50159999996, "line_count": 28},
]
yearly_total = 2016478.3214650005

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LLC TC Azimut 2026业绩"

header_fill = PatternFill("solid", fgColor="1F4E79")
even_fill   = PatternFill("solid", fgColor="EBF3FB")
total_fill  = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, row, col, value, bold=False, fill=None, fmt=None, align="center", color=None):
    c = ws.cell(row=row, column=col, value=value)
    fc = color if color else ("FFFFFF" if fill and fill.fgColor.rgb in ("1F4E79","2E75B6") else "000000")
    c.font = Font(bold=bold, color=fc)
    if fill:
        c.fill = fill
    c.border = border
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    return c

# 标题
ws.merge_cells("A1:C1")
t = ws["A1"]
t.value = 'LLC "TC Azimut"  2026年人民币业绩汇总'
t.font = Font(bold=True, size=14, color="1F4E79")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.append([])

# 表头
for col, h in enumerate(["月份", "人民币业绩（CNY）", "订单行数"], 1):
    cell(ws, 3, col, h, bold=True, fill=header_fill)
ws.row_dimensions[3].height = 20

# 数据行
for i, row in enumerate(monthly, 4):
    f = even_fill if i % 2 == 0 else None
    cell(ws, i, 1, row["month"], fill=f)
    cell(ws, i, 2, round(row["performance_cny"], 2), fill=f, fmt='#,##0.00')
    cell(ws, i, 3, row["line_count"], fill=f, fmt='#,##0')
    ws.row_dimensions[i].height = 18

# 合计行
total_row = 4 + len(monthly)
cell(ws, total_row, 1, "全年合计", bold=True, fill=total_fill)
cell(ws, total_row, 2, round(yearly_total, 2), bold=True, fill=total_fill, fmt='#,##0.00')
cell(ws, total_row, 3, sum(r["line_count"] for r in monthly), bold=True, fill=total_fill, fmt='#,##0')
ws.row_dimensions[total_row].height = 20

# 备注
note_row = total_row + 2
for r, txt in enumerate([
    "数据来源：sale_order_line_report（不限订单状态，按实际业绩字段汇总）",
    f"查询时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
    "注：5月含 partial_payment 状态订单，数据截至查询当日",
], note_row):
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(row=r, column=1, value=txt)
    c.font = Font(color="808080", italic=True, size=9)

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 14

stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = rf"D:\经销商PDCA\data_reports\{stamp}_LLC-TC-Azimut_2026业绩_v2.xlsx"
wb.save(out_path)
print(out_path)
