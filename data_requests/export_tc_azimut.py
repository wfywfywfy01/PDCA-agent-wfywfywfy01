import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

monthly = [
    {"month": "2026-01", "performance_cny": 618096.15432,  "line_count": 22},
    {"month": "2026-02", "performance_cny": 285708.3679799999, "line_count": 23},
    {"month": "2026-03", "performance_cny": 17236.38744,   "line_count": 9},
    {"month": "2026-04", "performance_cny": 580265.9101249999, "line_count": 40},
    {"month": "2026-05", "performance_cny": 24158.75,      "line_count": 8},
]
yearly_total = 1525465.569864999

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LLC TC Azimut 2026业绩"

header_fill   = PatternFill("solid", fgColor="1F4E79")
subtotal_fill = PatternFill("solid", fgColor="BDD7EE")
total_fill    = PatternFill("solid", fgColor="2E75B6")
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def styled_cell(ws, row, col, value, bold=False, fill=None, number_format=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color="FFFFFF" if fill and fill.fgColor.rgb in ("1F4E79","2E75B6") else "000000")
    if fill:
        c.fill = fill
    c.border = border
    c.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c

# 标题行
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = 'LLC "TC Azimut"  2026年人民币业绩汇总'
title_cell.font = Font(bold=True, size=14, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.append([])  # 空行

# 表头
headers = ["月份", "人民币业绩（CNY）", "订单行数"]
for col, h in enumerate(headers, 1):
    styled_cell(ws, 3, col, h, bold=True, fill=header_fill)

# 数据行
for i, row in enumerate(monthly, 4):
    styled_cell(ws, i, 1, row["month"])
    styled_cell(ws, i, 2, round(row["performance_cny"], 2), number_format='#,##0.00')
    styled_cell(ws, i, 3, row["line_count"], number_format='#,##0')
    if i % 2 == 0:
        for col in range(1, 4):
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

# 全年合计行
total_row = 4 + len(monthly)
styled_cell(ws, total_row, 1, "全年合计", bold=True, fill=total_fill)
styled_cell(ws, total_row, 2, round(yearly_total, 2), bold=True, fill=total_fill, number_format='#,##0.00')
styled_cell(ws, total_row, 3, sum(r["line_count"] for r in monthly), bold=True, fill=total_fill, number_format='#,##0')

# 备注
ws.cell(row=total_row + 2, column=1, value=f"数据来源：sale_order_line_report（状态：sale/done）")
ws.cell(row=total_row + 2, column=1).font = Font(color="808080", italic=True, size=9)
ws.merge_cells(f"A{total_row+2}:C{total_row+2}")
ws.cell(row=total_row + 3, column=1, value=f"查询时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}  /  5月数据截至查询当日，尚未完整")
ws.cell(row=total_row + 3, column=1).font = Font(color="808080", italic=True, size=9)
ws.merge_cells(f"A{total_row+3}:C{total_row+3}")

# 列宽
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 14

stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = rf"D:\经销商PDCA\data_reports\{stamp}_LLC-TC-Azimut_2026业绩.xlsx"
wb.save(out_path)
print(out_path)
