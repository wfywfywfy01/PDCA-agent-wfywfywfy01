# -*- coding: utf-8 -*-
"""导出海外门店周报表格（对齐截图格式）到桌面 Excel。"""
from __future__ import annotations

import json
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, PatternFill as PF

ROOT = Path(__file__).resolve().parents[1]
CONFIG = json.loads((ROOT / "config" / "dealers.json").read_text(encoding="utf-8"))
OVERVIEW = ROOT / "outputs" / "2026-W28_overview.json"
ACTIVATION_SCRIPT = (
    ROOT.parent / "data_platform" / "data_role_pdca_mvp" / "system_queries" / "dealer_activation_stats.py"
)
MONTHLY_SCRIPT = (
    ROOT.parent / "data_platform" / "data_role_pdca_mvp" / "system_queries" / "dealer_monthly_overseas.py"
)

# 周报重点跟踪门店（对齐截图顺序）
WATCHLIST = [
    "Billionaire Collections",
    "Safiranhamrah",
    "VERTU LONDON LTD",
    "My Shops Electronics Trading LLC",
    "VMG Communication and Technology Joint Stock Company",
    "Parth Kamlesh Doshi",
    "GURU ELECTRONICS SINGAPORE PTE LTD",
    "Luxem Store",
    "Dar Al Sabaek",
    "BIN BIN INVESTMENT(CAMBODIA) COLTD",
    "VST ECS (Thailand) Co., Ltd.",
    'LLC "TC Azimut"',
    "Bizcon Group",
    "Sidd Senthil",
    "ECN GmbH",
]

SERIES_ABBR = {
    "VERTU ALPHAFOLD": "AF",
    "VERTU AGENT Q": "AQ",
    "METAVERTU 2": "M2",
    "METAVERTU": "MV",
    "SIGNATURE S": "SIG S",
    "VERTU QUANTUM": "Quantum",
    "iVERTU": "iV",
    "METAWATCH S1": "M1",
    "VERTU PHANTOM": "Phantom",
    "VERTU IRONFLIP": "IF",
    "VERTU METAVERTU 2": "M2",
}


def vertu_cmd() -> list[str]:
    npm = Path.home() / "AppData" / "Roaming" / "npm"
    cjs = npm / "node_modules" / "@vertu-tech" / "vps-cli" / "dist" / "vertu.cjs"
    if cjs.exists():
        return ["node", str(cjs)]
    cmd = npm / "vertu.cmd"
    return [str(cmd)] if cmd.exists() else ["vertu"]


def run_sandbox(code_path: Path, params: dict | None = None) -> dict:
    cmd = [*vertu_cmd(), "odoo", "data", "sandbox", "--code-file", str(code_path)]
    if params:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8") as f:
            json.dump(params, f)
            cmd += ["--params-file", f.name]
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr or proc.stdout)
    data = json.loads(proc.stdout.strip())
    ex = data["result"]["execution"]
    if ex.get("error"):
        raise RuntimeError(json.dumps(ex["error"], ensure_ascii=False))
    return ex["result"]


def match_dealer(partner: str) -> str | None:
    p = partner or ""
    for d in CONFIG["dealers"]:
        for m in d["match"]:
            if m.lower() in p.lower():
                return d["name"]
    return None


def fmt_series(series: str) -> str:
    s = re.sub(r"<br\s*/?>", " ", series or "").strip()
    for full, abbr in SERIES_ABBR.items():
        if full.upper() in s.upper():
            return abbr
    if len(s) > 18:
        return s[:16] + "…"
    return s


def fmt_products(items: list[dict], limit: int = 3) -> str:
    parts = []
    for it in items[:limit]:
        abbr = fmt_series(it["series"])
        qty = int(it["qty"]) if it["qty"] == int(it["qty"]) else it["qty"]
        wan = it["amount"] / 10000
        parts.append(f"{abbr}×{qty} (¥{wan:.2f}万)")
    return "，".join(parts)


def owner_label(sales: str) -> str:
    if sales in ("杨晶晶", "何海文"):
        return "杨晶晶/何海文"
    if sales == "于冰":
        return "于冰"
    if sales in ("Lina", "Viki"):
        return sales
    return sales


def remark(row: dict, products: list[dict]) -> str:
    si = row["amount"]
    qty = row["qty"]
    act = row["activated"]
    if si <= 0 and qty > 0:
        return f"销量{int(qty)}台但VPS SI未录入，待补录"
    if si <= 0:
        return "本周无SI录入"
    majors = {fmt_series(p["series"]) for p in products}
    non_phone = any(
        p["series"] in ("权益服务", "预售虚拟类", "其他") or p["amount"] > 0 and p.get("major") == "其他"
        for p in products
    )
    phone_qty = sum(p["qty"] for p in products if p.get("major") == "手机")
    if qty >= 20 and phone_qty <= 3:
        return f"销量{int(qty)}件，多为非手机（辅料/IOT/权益等），不代表手机动销"
    if act > 0 and si <= 0:
        return f"激活{act}台，SI待补录"
    if si > 0 and act == 0 and qty > 0:
        return "有SI/销量，激活待跟进"
    if si > 100000:
        return "维持，补录五件套SO套表"
    return ""


def todo(row: dict, sales: str) -> str:
    si = row["amount"]
    qty = row["qty"]
    act = row["activated"]
    owner = sales if sales not in ("杨晶晶", "何海文") else "杨晶晶"
    if si <= 0 and qty > 0:
        return f"催补SI录入（{owner}）"
    if si <= 0 and act <= 0:
        return f"催补SI录入（{owner}）"
    if act <= 0 and qty > 0:
        return "补录激活订单"
    if si > 100000:
        return "维持，补录五件套SO套表"
    if "VMG" in row["name"]:
        return "AF高客单尾款 + 4店巡店（于冰）"
    if row["name"] == "Safiranhamrah":
        return "SI录入 + 选品（$30k+$20k USD）（Viki）"
    if row["name"] == "My Shops Electronics Trading LLC":
        return "MYGROUP紧急跟进（Lina）"
    return f"维持跟进（{owner}）"


def build_rows(mtd_map: dict, prod_map: dict, act_map: dict) -> list[dict]:
    meta = {d["name"]: d for d in CONFIG["dealers"]}
    total_si = sum(v["amount"] for v in mtd_map.values())
    rows = []
    for name in WATCHLIST:
        d = mtd_map.get(name, {"amount": 0, "lines": 0, "qty": 0, "orders": 0})
        m = meta.get(name, {"country": "—", "sales": "—"})
        prods = prod_map.get(name, [])
        act = act_map.get(name, 0)
        amt = d["amount"]
        qty = d.get("qty") or d.get("lines") or 0
        pct = round(amt / total_si * 100, 1) if total_si and amt > 0 else 0
        rows.append({
            "name": name,
            "country": m.get("country", "—"),
            "owner": owner_label(m.get("sales", "—")),
            "amount": amt,
            "qty": qty,
            "orders": d.get("orders", 0),
            "activated": act,
            "products": prods,
            "pct": pct,
            "remark": remark({"name": name, "amount": amt, "qty": qty, "activated": act}, prods),
            "todo": todo({"name": name, "amount": amt, "qty": qty, "activated": act}, m.get("sales", "")),
        })
    rows.sort(key=lambda x: (-x["amount"], -x["qty"]))
    return rows, total_si


def export_excel(rows: list[dict], total_si: float, period: str, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "门店周报"
    ws["A1"] = f"7月门店数据 ({period})"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    headers = [
        "门店", "国家", "Owner", "SI(万)", "销量", "订单", "激活",
        "主力货品", "占比", "备注", "下周待办",
    ]
    ws.append(headers)
    fill = PF("solid", fgColor="1E3A5F")
    hf = Font(bold=True, color="FFFFFF", size=10)
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill, cell.font, cell.alignment = fill, hf, ac

    for r in rows:
        ws.append([
            r["name"],
            r["country"],
            r["owner"],
            round(r["amount"] / 10000, 2) if r["amount"] else 0,
            int(r["qty"]),
            r["orders"],
            r["activated"],
            fmt_products(r["products"]),
            f"{r['pct']}%" if r["pct"] else "—",
            r["remark"],
            r["todo"],
        ])

    tot_qty = sum(r["qty"] for r in rows)
    tot_act = sum(r["activated"] for r in rows)
    ws.append([
        "合计", "", "",
        round(total_si / 10000, 2) if total_si else 0,
        int(tot_qty), "", tot_act, "", "", "", "",
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # 门店总结
    ws.append([])
    ws.append(["门店总结"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    highlights = [r for r in rows if r["amount"] >= 150000 or (r["qty"] >= 9 and r["amount"] <= 0)]
    for r in highlights[:6]:
        line = f"{r['name']}：SI ¥{r['amount']/10000:.2f}万"
        if r["activated"]:
            line += f"，激活{r['activated']}台"
        if r["products"]:
            line += "，" + fmt_products(r["products"], 2)
        if r["remark"]:
            line += " → " + r["remark"]
        ws.append([line])

    widths = [36, 10, 14, 8, 6, 6, 6, 42, 6, 28, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(out_path)


def main() -> int:
    overview = json.loads(OVERVIEW.read_text(encoding="utf-8"))
    period = f"{overview['meta']['mtd_start']} ~ {overview['meta']['mtd_end']}"

    # 重新拉 sandbox 原始 mtd 经销商 + 产品
    sys.path.insert(0, str(ROOT / "scripts"))
    from fetch_overview import build_case_sql, resolve_window, run_vertu_sandbox, write_sandbox  # noqa: WPS433

    window = resolve_window(type("A", (), {"as_of": overview["meta"]["as_of"], "week_start": "", "week_end": ""})())
    write_sandbox(CONFIG["dealers"], window)
    raw = run_vertu_sandbox()

    mtd_map: dict[str, dict] = {}
    for r in raw.get("mtd", []):
        mtd_map[r["dealer"]] = {
            "amount": float(r["amount"] or 0),
            "lines": int(r["lines"] or 0),
            "qty": int(r["lines"] or 0),
            "orders": 0,
        }

    # 补 qty / orders
    sandbox_store = ROOT / "scripts" / "_sandbox_store_weekly_mtd.py"
    sandbox_store.write_text(_mtd_extra_sql(build_case_sql(CONFIG["dealers"]), window["mtd_start"], window["mtd_end"]), encoding="utf-8")
    extra = run_sandbox(sandbox_store)
    for r in extra.get("summary", []):
        name = r["dealer"]
        if name in mtd_map:
            mtd_map[name]["qty"] = float(r.get("qty") or 0)
            mtd_map[name]["orders"] = int(r.get("orders") or 0)

    prod_map: dict[str, list] = defaultdict(list)
    for r in raw.get("products_dealer_mtd", []):
        prod_map[r["dealer"]].append({
            "series": r["series"],
            "major": r["major"],
            "qty": r["lines"],
            "amount": float(r["amount"] or 0),
        })
    for lst in prod_map.values():
        lst.sort(key=lambda x: -x["amount"])

    # 激活（月内）
    act_raw = run_sandbox(
        MONTHLY_SCRIPT,
        {"run_date": window["mtd_end"], "start_date": window["mtd_start"], "end_date": window["mtd_end"]},
    )
    act_map: dict[str, int] = {}
    for r in act_raw.get("dealers", []):
        mapped = match_dealer(r.get("dealer_name", "")) or r.get("dealer_name")
        if mapped:
            act_map[mapped] = act_map.get(mapped, 0) + int(r.get("month_activated") or 0)

    rows, total_si = build_rows(mtd_map, prod_map, act_map)
    today = date.today().strftime("%m%d")
    out = Path.home() / "Desktop" / f"海外门店周报_7月_{today}.xlsx"
    export_excel(rows, total_si, period, out)
    print(str(out))
    print(f"period={period} stores={len(rows)} total_si_wan={total_si/10000:.2f}")
    for r in rows[:8]:
        print(f"  {r['name'][:30]:30} SI={r['amount']/10000:6.2f}万 qty={int(r['qty']):3} act={r['activated']}")
    return 0


def _mtd_extra_sql(case_sql: str, start: str, end: str) -> str:
    return f'''# mtd qty/orders
CASE_SQL = """{case_sql}"""
sql = (
    "SELECT (" + CASE_SQL + ") AS dealer_name, "
    "SUM(COALESCE(\\\"数量\\\", 0)) AS qty, "
    "COUNT(DISTINCT \\\"原订单号\\\") AS orders "
    "FROM odoo_sale WHERE \\\"销售日期\\\" >= '{start}' "
    "AND \\\"销售日期\\\" <= '{end} 23:59:59' "
    "AND (" + CASE_SQL + ") IS NOT NULL GROUP BY 1"
)
rows = sql_read(sql)
ai["result"] = {{"summary": [{{"dealer": r.get("dealer_name"), "qty": float(r.get("qty") or 0), "orders": int(r.get("orders") or 0)}} for r in rows]}}
'''


if __name__ == "__main__":
    raise SystemExit(main())
