# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path

import openpyxl

p = Path(r"c:\Users\frank\Downloads\代理商终销Distribution Sell out.xlsx")
out = Path(__file__).resolve().parents[1] / "modules" / "walkin_cockpit" / "data" / "_dealer_parse_preview.json"

REGION_MAP = {
    "中国": "东区",
    "香港": "南区",
    "澳门": "南区",
    "台湾": "东区",
    "新加坡": "南区",
    "马来西亚": "南区",
    "泰国": "越南区",
    "越南": "越南区",
    "印尼": "越南区",
    "印度尼西亚": "越南区",
    "菲律宾": "越南区",
    "巴基斯坦": "西区",
    "哈萨克斯坦": "西区",
    "乌兹别克斯坦": "西区",
    "俄罗斯": "北区",
    "土耳其": "西区",
    "阿联酋": "西区",
    "沙特": "西区",
    "卡塔尔": "西区",
    "科威特": "西区",
    "英国": "北区",
    "法国": "北区",
    "德国": "北区",
    "意大利": "北区",
    "西班牙": "北区",
    "美国": "北区",
    "加拿大": "北区",
}


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_country(c):
    if not c:
        return ""
    return str(c).strip()


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = None
    dealers = []
    current_team = ""
    for row in rows:
        cells = list(row) + [None] * 20
        if cells[0] and "Team" in str(cells[0]):
            header = [str(c or "").replace("\n", " ") for c in row]
            continue
        if not header:
            continue
        if cells[0]:
            current_team = str(cells[0]).strip()
        country = norm_country(cells[1])
        dealer = cells[5] if len(cells) > 5 else None
        if not dealer:
            continue
        dealer = str(dealer).strip()
        if not dealer or dealer in ("Dealer Name", "代理商名称"):
            continue
        amount = _num(cells[14]) if len(cells) > 14 else 0
        qty = _num(cells[13]) if len(cells) > 13 else 0
        ctype = str(cells[3] or "").strip() if len(cells) > 3 else ""
        dealers.append({
            "team": current_team,
            "country": country,
            "dealerName": dealer,
            "customerType": ctype,
            "sellOutQty": qty,
            "amount": amount,
        })
    return dealers


def main():
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    all_rows = []
    for name in wb.sheetnames:
        all_rows.extend(parse_sheet(wb[name]))
    wb.close()
    by_team = {}
    for d in all_rows:
        by_team.setdefault(d["team"], []).append(d)
    preview = {
        "total": len(all_rows),
        "teams": {k: len(v) for k, v in by_team.items()},
        "sample": all_rows[:25],
        "top_amount": sorted(all_rows, key=lambda x: x["amount"], reverse=True)[:15],
    }
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")
    print("written", out, "total", len(all_rows))


if __name__ == "__main__":
    main()
