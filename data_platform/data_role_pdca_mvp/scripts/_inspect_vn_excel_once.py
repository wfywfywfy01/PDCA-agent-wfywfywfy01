# -*- coding: utf-8 -*-
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

p = Path(r"c:\Users\frank\Desktop\Data collecet(5).xlsx")
out = Path(__file__).resolve().parents[1] / "modules" / "walkin_cockpit" / "data" / "_excel_inspect.json"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
report = {"sheets": wb.sheetnames, "samples": {}}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        rows.append([str(c) if c is not None else None for c in (row or ())])
    report["samples"][name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows": rows,
    }
wb.close()
out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print("wrote", out)
