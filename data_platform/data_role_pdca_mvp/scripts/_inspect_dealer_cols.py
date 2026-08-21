# -*- coding: utf-8 -*-
import json
from pathlib import Path
import openpyxl

p = Path(r"c:\Users\frank\Downloads\代理商终销Distribution Sell out.xlsx")
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
result = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    result[name] = [
        {j: (str(v)[:50] if v is not None else None) for j, v in enumerate(row[:22])}
        for i, row in enumerate(rows)
    ]
    # sample data row with max non-null
    data_rows = list(ws.iter_rows(min_row=3, max_row=12, values_only=True))
    result[name + "_data"] = [
        {j: v for j, v in enumerate(row[:22]) if v is not None}
        for row in data_rows[:5]
    ]
wb.close()
out = Path(__file__).resolve().parents[1] / "modules" / "walkin_cockpit" / "data" / "_dealer_cols.json"
out.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
