# -*- coding: utf-8 -*-
# 由 pdca_workbench.py 按域拆分生成：销售数据查询与研究类查询
# 本文件不单独 import：由 pdca_workbench/__init__.py 以共享命名空间按原顺序 exec，
# 与拆分前单文件语义完全一致。所有符号请通过 `import pdca_workbench` 访问。


def sales_aliases():
    path = WORKSPACE / "config" / "sales_aliases.csv"
    aliases = {}
    if not path.exists():
        return aliases
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                raw = (row.get("raw_sales") or "").strip()
                canonical = (row.get("canonical_sales") or "").strip()
                if raw and canonical:
                    aliases[raw.lower()] = canonical
                    aliases[canonical.lower()] = canonical
    except Exception:
        return aliases
    return aliases


def canonical_sales_name(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return sales_aliases().get(text.lower(), text)


def sales_data_files():
    raw_dir = REPO_ROOT / "data_raw"
    if not raw_dir.exists():
        return []
    return sorted(
        raw_dir.glob("dealer_sales_month_to_date_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def sales_data_file_for_range(run_date, start_date=None):
    suffix = run_date if not start_date or start_date == run_date else f"{start_date}_to_{run_date}"
    return REPO_ROOT / "data_raw" / f"dealer_sales_month_to_date_{suffix}.json"


def is_fresh_sales_data(path):
    if not path or not Path(path).exists():
        return False
    age_seconds = datetime.now().timestamp() - Path(path).stat().st_mtime
    if age_seconds > RAW_SALES_CACHE_SECONDS:
        return False
    try:
        payload = read_json(Path(path))
    except Exception:
        return False
    result = None
    if isinstance(payload, dict):
        if isinstance(payload.get("execution"), dict):
            result = payload["execution"].get("result")
        if result is None and isinstance(payload.get("ai"), dict):
            result = payload["ai"].get("result")
        if result is None:
            result = payload.get("result")
    return isinstance(result, dict) and bool(result.get("summary_mode"))


def latest_sales_data_payload():
    for path in sales_data_files():
        try:
            payload = read_json(path)
        except Exception:
            continue
        result = None
        if isinstance(payload, dict):
            if isinstance(payload.get("execution"), dict):
                result = payload["execution"].get("result")
            if result is None and isinstance(payload.get("ai"), dict):
                result = payload["ai"].get("result")
            if result is None:
                result = payload.get("result")
        if isinstance(result, dict) and result.get("summary_mode"):
            return path, result
    return None, None


def is_sales_data_query(query):
    text = query.lower()
    if is_logistics_query(query):
        return False
    intent_words = ["业绩", "销售", "出表", "excel", "表格", "拉一下", "拉取", "客户", "产品", "团队"]
    return any(word in text for word in intent_words)


def is_research_query(query):
    text = str(query or "").lower()
    keywords = [
        "调研", "研究", "竞品", "市场", "客户背景", "背景调查", "公开资料", "行业",
        "渠道", "国家", "政策", "资料整理", "分析一下", "research", "market", "competitor",
    ]
    return any(keyword in text for keyword in keywords)


def is_vps_cli_query(query):
    return "从vps" in str(query or "").lower().replace(" ", "")


def requested_vps_date_range(query):
    text = str(query or "")
    today = datetime.now()
    iso_match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if iso_match:
        year, month, day = map(int, iso_match.groups())
        run_date = datetime(year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    md_match = re.search(r"(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?", text)
    if md_match:
        month, day = map(int, md_match.groups())
        run_date = datetime(today.year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    month_match = re.search(r"(\d{1,2})\s*月|([一二三四五六七八九十])月", text)
    if month_match:
        chinese_months = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
        month = int(month_match.group(1)) if month_match.group(1) else chinese_months.get(month_match.group(2), today.month)
        day = today.day if month == today.month else 28
        run_date = datetime(today.year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    return today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-01")


def query_requires_fresh_vps(query):
    text = str(query or "").lower()
    return any(word in text for word in ["刷新", "重新", "最新", "实时", "强制"])


def pull_vps_sales_data(run_date, start_date=None, force=False):
    cached_path = sales_data_file_for_range(run_date, start_date)
    if not force and is_fresh_sales_data(cached_path):
        return cached_path
    puller = WORKSPACE / "scripts" / "pull_vps_sales_data.ps1"
    if not puller.exists():
        raise RuntimeError(f"VPS 拉数脚本不存在：{puller}")
    command = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(puller),
        "-Date",
        run_date,
        "-Workspace",
        str(WORKSPACE),
    ]
    if start_date:
        command.extend(["-StartDate", start_date])
    try:
        completed = subprocess.run(
            command,
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        if cached_path.exists():
            return cached_path
        raise RuntimeError("VPS-CLI 拉数超过 120 秒，已自动终止。")
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
    if completed.returncode != 0:
        raise RuntimeError(output or "VPS-CLI 拉数失败")
    for line in reversed(output.splitlines()):
        path = Path(line.strip().strip('"'))
        if path.exists():
            return path
    fallback = sales_data_file_for_range(run_date, start_date)
    return fallback if fallback.exists() else None


def run_vps_cli_query(query):
    run_date, start_date = requested_vps_date_range(query)
    try:
        if is_sales_data_query(query):
            pulled_path = pull_vps_sales_data(run_date, start_date, force=query_requires_fresh_vps(query))
            result = run_sales_data_query(query)
            if result.get("ok"):
                result["content"] = (
                    "已按「从vps」规则调用 VPS-CLI 拉取最新数据，并直接生成结果文件。\n\n"
                    f"- VPS 原始数据：{Path(pulled_path).name if pulled_path else '已刷新'}\n"
                    + result.get("content", "")
                )
                result["kind"] = "vps_sales_excel"
            return result
    except Exception as exc:
        return {"ok": False, "content": f"VPS-CLI 执行失败：{exc}", "path": None, "kind": "vps_error"}

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = DATA_REPORTS / f"{stamp}_vps-cli-routing_summary.md"
    content = "\n".join([
        "# VPS-CLI 查询路由结果",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 用户问题",
        "",
        query,
        "",
        "## 处理结果",
        "",
        "已识别到关键词「从vps」，但当前工作台只内置了经销商业绩类查询到 Excel 的稳定映射。",
        "",
        "## 已支持",
        "",
        "- 从vps拉销售/业绩/产品/客户/团队汇总：自动调用 VPS-CLI，生成 Excel。",
        "- 从vps拉某个销售（例如 Lina）的五月业绩：自动筛选销售员并生成 Excel。",
        "",
        "## 下一步",
        "",
        "请把需求写成类似：`从vps拉一下lina五月业绩`、`从vps拉五月经销商业绩表`。",
    ])
    write_text(path, content)
    return {"ok": True, "content": content, "path": str(path), "filename": path.name, "kind": "vps_markdown"}


def run_research_chat(query):
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    topic = "research-agent"
    out_path = DATA_REPORTS / f"{stamp}_{topic}_summary.md"
    prompt = f"""
你是经销商 PDCA 工作台的 research-agent，负责公开资料调研。

用户调研问题：
{query}

请遵守：
1. 只使用公开资料和用户提供资料。
2. 不编造来源；没有可靠来源时写“未找到可靠公开来源”。
3. 每条关键结论尽量附来源 URL 或来源名称。
4. 最多 15 步，超过即停止并说明当前进展。
5. 输出不超过 4000 tokens。
6. 不登录第三方账号，不绕过验证码，不访问需要付费或权限的内容。

输出 Markdown，必须包含：
# 调研报告
## 调研问题
## 核心结论
## 关键事实
## 证据来源
## 不确定事项
## 建议下一步
""".strip()
    command = [
        hermes_exe(),
        "chat",
        "-q",
        prompt,
        "-Q",
        "--max-turns",
        "15",
    ]
    try:
        completed = subprocess.run(
            command,
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "content": "调研 Agent 执行超过 120 秒，已强制终止。请缩小调研范围后重试。", "path": None, "kind": "research_timeout"}
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
    if completed.returncode != 0 and not output:
        return {"ok": False, "content": "调研 Agent 调用失败，且没有返回错误信息。", "path": None, "kind": "research_error"}
    content = output or "调研 Agent 已运行，但没有返回内容。"
    if "Error code:" in content or "Incorrect API key" in content:
        return {"ok": False, "content": f"调研 Agent 调用失败：{content}", "path": None, "kind": "research_error"}
    report = "\n\n".join([
        content,
        "",
        f"---\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\nAgent: research-agent",
    ])
    write_text(out_path, report)
    return {"ok": True, "content": report[-4000:], "path": str(out_path), "filename": out_path.name, "kind": "research"}


def requested_salesperson(query, rows):
    text = query.lower()
    aliases = sales_aliases()
    candidates = set(aliases.values())
    candidates.update(canonical_sales_name(row.get("salesperson")) for row in rows or [])
    for name in sorted((item for item in candidates if item), key=len, reverse=True):
        if name.lower() in text:
            return name
    if "lina" in text:
        return "Lina"
    return ""


def summary_row(row, key_name):
    return [
        canonical_sales_name(row.get(key_name)) if key_name == "salesperson" else (row.get(key_name) or ""),
        float(row.get("performance") or 0),
        float(row.get("quantity") or 0),
        int(float(row.get("line_count") or row.get("rows") or 0)),
    ]


def run_sales_data_query(query):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        return {"ok": False, "content": f"本机缺少 openpyxl，无法生成 Excel：{exc}", "path": None}

    data_path, result = latest_sales_data_payload()
    if not result:
        return {"ok": False, "content": "没有找到可用的 VPS 月累业绩数据，请先运行数据拉取。", "path": None}

    salesperson_rows = result.get("salesperson_summary") or []
    target_sales = requested_salesperson(query, salesperson_rows)
    filtered_sales = [
        row for row in salesperson_rows
        if not target_sales or canonical_sales_name(row.get("salesperson")).lower() == target_sales.lower()
    ]
    if target_sales and not filtered_sales:
        return {"ok": False, "content": f"已读取 {data_path.name}，但没有找到销售员 {target_sales} 的业绩。", "path": None}

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_target = safe_name(target_sales or "all-sales")
    out_path = DATA_REPORTS / f"{stamp}_sales_performance_{safe_target}.xlsx"
    DATA_REPORTS.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title, rows, key_name):
        ws = wb.create_sheet(title[:31])
        headers = ["维度", "业绩", "数量", "明细行数"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F62BD")
        for row in rows or []:
            ws.append(summary_row(row, key_name))
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    add_sheet("销售员汇总", filtered_sales or salesperson_rows, "salesperson")
    add_sheet("团队汇总", result.get("team_summary") or [], "team")
    add_sheet("产品TOP", result.get("product_summary") or [], "product_name")
    add_sheet("客户TOP", result.get("customer_summary") or [], "partner_name")
    wb.save(out_path)

    total_perf = sum(float(row.get("performance") or 0) for row in (filtered_sales or salesperson_rows))
    total_qty = sum(float(row.get("quantity") or 0) for row in (filtered_sales or salesperson_rows))
    label = target_sales or "全部销售员"
    content = (
        f"已直接从 VPS 月累数据生成 Excel，不再走 Hermes 对话。\n\n"
        f"- 查询对象：{label}\n"
        f"- 数据文件：{data_path.name}\n"
        f"- 数据周期：{result.get('month_start', '')} 至 {result.get('run_date', '')}\n"
        f"- 业绩合计：{total_perf:,.2f}\n"
        f"- 数量合计：{total_qty:,.2f}\n"
        f"- 输出文件：{out_path.name}"
    )
    return {"ok": True, "content": content, "path": str(out_path), "filename": out_path.name, "kind": "sales_excel"}
