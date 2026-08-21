"""解析货代预报 xlsx，按顺丰单号折叠品名行。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from db import upsert

CHANNEL_TO_CARRIER = {
    "深圳UPS": "UPS",
    "香港DHL": "DHL",
    "专线": "专线",
    "港车": "港车",
}

PHONE_KEYS = ("手机", "cellphone", "mobile phone", "iphone")
WATCH_KEYS = ("手表", "watch")


PREFIX = "Transfer Information Lines/"

# 日升预报标准表头（samples/日升预报.xlsx）。带前缀和裸列名都能取。
FORECAST_FIELDS = (
    "日期",
    "客户单号",
    "顺丰单号",
    "转运单号",
    "渠道",
    "收货人",
    "收件人公司",
    "收件人英文国家",
    "收件人地址1",
    "州",
    "城市",
    "邮编",
    "电话",
    "英文名",
    "中文名称",
    "产品",
    "数量",
    "重量",
    "单价(USD)",
    "总价(USD)",
    "材质",
    "用途",
    "HS Code",
    "邮箱",
    "备注",
)


def _clean(val) -> str:
    """单元格转字符串。斜杠、None 当空。
    @param {*} val
    @returns {str}
    """
    if val is None:
        return ""
    text = str(val).replace("\u200b", "").strip()
    if text in ("/", "None", "nan"):
        return ""
    return text


def _is_junk_row(raw) -> bool:
    """全空或纯斜杠分隔行。
    @param {tuple|list} raw
    @returns {bool}
    """
    values = [_clean(v) for v in raw]
    if not any(values):
        return True
    first = (values[0] or "").replace("/", "").replace(",", "").replace("，", "").replace(" ", "")
    return first == "" and not any(values[1:])


def _header_row(ws):
    """定位含客户单号的表头，最多扫 20 行。
    @param {*} ws
    @returns {tuple} (行号, 表头列表)
    """
    for i, raw in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        blob = " ".join(str(c or "") for c in raw)
        if "客户单号" in blob or "Transfer Information Lines" in blob:
            return i, list(raw)
    first = next(ws.iter_rows(min_row=1, max_row=1))
    return 1, [c.value for c in first]


def _v(row: dict, name: str):
    """取预报列。兼容 Transfer Information Lines/ 前缀。
    @param {dict} row
    @param {str} name
    """
    val = row.get(f"{PREFIX}{name}", row.get(name))
    return val


def _fmt_date(val) -> str | None:
    if val is None or val == "/" or val == "":
        return None
    if isinstance(val, datetime):
        return f"{val.year}/{val.month}/{val.day}"
    text = str(val).strip()
    return text or None


def _num(val) -> float:
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0


def _is_phone(en: str, zh: str) -> bool:
    blob = f"{en} {zh}".lower()
    return any(k in blob for k in PHONE_KEYS)


def _is_watch(en: str, zh: str) -> bool:
    blob = f"{en} {zh}".lower()
    return any(k in blob for k in WATCH_KEYS)


def parse_forecast(path: str | Path) -> list[dict]:
    """读预报，返回按顺丰单号折叠后的行。
    @param {str} path
    @returns {list}
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_i, headers = _header_row(ws)
    groups: list[dict] = []
    current = None
    for raw in ws.iter_rows(min_row=header_i + 1, values_only=True):
        if _is_junk_row(raw):
            continue
        row = dict(zip(headers, raw))
        sf = _clean(_v(row, "顺丰单号"))
        if sf:
            current = {
                "_sf": sf,
                "_order": _clean(_v(row, "客户单号")),
                "_intl": _clean(_v(row, "转运单号")),
                "_channel": _clean(_v(row, "渠道")),
                "_recv": _clean(_v(row, "收货人")),
                "_country": _clean(_v(row, "收件人英文国家")),
                "_date": _v(row, "日期"),
                "_notes": [],
                "_items": [],
            }
            note = _clean(_v(row, "备注"))
            if note:
                current["_notes"].append(note)
            extra = _clean(row.get("备注"))
            if extra and extra not in current["_notes"]:
                current["_notes"].append(extra)
            groups.append(current)
        if current is None:
            continue
        order = _clean(_v(row, "客户单号"))
        if order:
            current["_order"] = order
        recv = _clean(_v(row, "收货人"))
        if recv:
            current["_recv"] = recv
        country = _clean(_v(row, "收件人英文国家"))
        if country:
            current["_country"] = country
        channel = _clean(_v(row, "渠道"))
        if channel:
            current["_channel"] = channel
        en = _clean(_v(row, "英文名"))
        zh = _clean(_v(row, "中文名称"))
        if not en and not zh:
            continue
        current["_items"].append(
            {
                "en": en,
                "zh": zh,
                "qty": _num(_v(row, "数量")),
                "weight": _num(_v(row, "重量")),
            }
        )
        note = _clean(_v(row, "备注"))
        if note and note not in current["_notes"]:
            current["_notes"].append(note)
    out = []
    for g in groups:
        names = []
        qty = 0.0
        weight = 0.0
        phones = 0.0
        watches = 0.0
        for it in g["_items"]:
            label = it["zh"] or it["en"]
            if it["en"] and it["zh"] and it["en"] != it["zh"]:
                label = f"{it['zh']}({it['en']})"
            names.append(label)
            qty += it["qty"]
            weight += it["weight"]
            if _is_phone(it["en"], it["zh"]):
                phones += it["qty"]
            if _is_watch(it["en"], it["zh"]):
                watches += it["qty"]
        ship_date = _fmt_date(g["_date"])
        month = ""
        if isinstance(g["_date"], datetime):
            month = str(g["_date"].month)
        elif ship_date and "/" in ship_date:
            parts = ship_date.split("/")
            month = parts[1] if len(parts) > 1 else ""
        carrier = CHANNEL_TO_CARRIER.get(g["_channel"], g["_channel"])
        intl = g["_intl"]
        rec = {
            "订单号": g["_order"],
            "共建销售单号/系统单号": g["_order"],
            "顺丰单号": g["_sf"],
            "快递公司": carrier,
            "境外收货人": g["_recv"],
            "目的地": g["_country"],
            "产品名称": ";".join(names),
            "货品总数": str(int(qty) if qty == int(qty) else qty),
            "手机台数": str(int(phones)),
            "手表台": str(int(watches)),
            "计费重量KG": str(weight),
            "件数/箱数": "1",
            "发货日期": ship_date,
            "月份": month,
            "是否发预报（报关资料）": "是",
            "签收状态": "已出国际单" if intl else "已预报",
            "备注": ";".join(g["_notes"]),
            "预报文件": str(path),
        }
        if intl:
            rec["国际单号"] = intl
        out.append(rec)
    return out


FORECAST_OVERWRITE = (
    "订单号",
    "共建销售单号/系统单号",
    "快递公司",
    "境外收货人",
    "目的地",
    "产品名称",
    "货品总数",
    "手机台数",
    "手表台",
    "计费重量KG",
    "发货日期",
    "月份",
    "是否发预报（报关资料）",
    "预报文件",
    "备注",
    "件数/箱数",
    "国际单号",
)


def ingest_forecast(path: str | Path) -> int:
    """解析并写入 sqlite。生命周期只向前，不把已出国际单打回已预报。
    @param {str} path
    @returns {int} 票数
    """
    from db import get
    from status import advance, display_status, lifecycle_of

    rows = parse_forecast(path)
    for row in rows:
        upsert(row, overwrite=False)
        patch = {k: row[k] for k in FORECAST_OVERWRITE if row.get(k)}
        patch["顺丰单号"] = row["顺丰单号"]
        upsert(patch, overwrite=True)
        saved = get(row["顺丰单号"]) or {}
        target = "LABELED" if saved.get("国际单号") else "PENDING"
        life = advance(lifecycle_of(saved), target)
        exc = saved.get("异常") or ""
        if life == "PENDING" and (
            saved.get("签收状态") == "待人工" or (saved.get("匹配级别") or "").upper() == "C"
        ):
            shown = "待人工"
        else:
            shown = display_status(life, exc)
        upsert(
            {
                "顺丰单号": row["顺丰单号"],
                "生命周期": life,
                "签收状态": shown,
            },
            overwrite=True,
        )
    return len(rows)
